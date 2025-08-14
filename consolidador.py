# SEÇÃO DE IMPORTAÇÕES DE BIBLIOTECAS

import pandas as pd
import openpyxl
import os
import shutil
import sys
import tkinter
from tkinter import filedialog, messagebox
import logging
import traceback
import threading
import queue
from datetime import datetime
from dateutil.parser import parse
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import customtkinter as ctk
from PIL import Image

# CONFIGURAÇÃO DO LOGGING
logging.basicConfig(filename='consolidador.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s', encoding='utf-8')

# FUNÇÕES AUXILIARES

def resource_path(relative_path): # Função para obter o caminho do arquivo
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def encontrar_linha_inicial_real(sheet, header_rows=2): # Função para encontrar a linha inicial real
    for r in range(sheet.max_row, header_rows, -1):
        if sheet.cell(row=r, column=3).value:
            return r + 1
    return header_rows + 1

def clean_number_value(value): # Função para limpar o valor do número
    s = str(value).strip()
    if ',' in s:
        return s.replace('.', '').replace(',', '.')
    return "".join(c for c in s if c.isdigit() or c == '.')

# NÚCLEO LÓGICO (O "MOTOR" DO APLICATIVO)   

def processar_arquivo_origem(arquivo, colunas_mestra, log_queue): # Função para processar o arquivo de origem
    def log(msg):
        log_queue.put(msg)
    
    try:
        workbook = openpyxl.load_workbook(arquivo, data_only=True)
        sheet = workbook.worksheets[0]
        number, seller = sheet['C3'].value, sheet['C8'].value
        
        if not number or not seller:
            log(f"  -> ERRO: Não foi possível extrair NUMBER(C3) ou SELLER(C8) de {os.path.basename(arquivo)}.")
            return None, None

        ref_cucala, date_raw, buyer, agent, n_ag, n_buyer = (
            sheet['C13'].value, sheet['C4'].value, sheet['C9'].value,
            sheet['C10'].value, sheet['E9'].value, sheet['E10'].value
        )
        date_formatada = date_raw
        if date_raw:
            try: date_formatada = parse(str(date_raw)).strftime('%d/%m/%Y')
            except Exception: pass
        
        df_dados = pd.read_excel(arquivo, engine='openpyxl', skiprows=13)
        mapa_de_traducao = {
            'ORIGEM': 'ORIGIN', 'MUNICIPIO': 'GIN LOCATION', 'FAZENDA': 'FAZENDA(FARM NAME)',
            'FAZENDA (FARM NAME)': 'FAZENDA(FARM NAME)', 'LOTE': 'LOT NO.', 'FARDOS': 'BALES',
            'P.LIQUIDO': 'Net Weight', 'TARA': 'Tare', 'P.BRUTO': 'GROSS KG', 'TIPO': 'GRADE',
            'FIBRA': 'STAPLE', 'FOLHA': 'LEAF', 'COR': 'COLOR', 'BENEFICIO': 'CHARACTER',
            'TIPO DO VENDEDOR': 'TYPE AGREED', 'OBSERVAÇÃO': 'P&D'
        }
        df_dados.columns = df_dados.columns.str.strip()
        df_dados.rename(columns=mapa_de_traducao, inplace=True)

        colunas_para_limpar = ['BALES', 'Net Weight', 'GROSS KG', 'MIC.', 'GPT', 'UHM']
        for col in colunas_para_limpar:
            if col in df_dados.columns:
                df_dados[col] = pd.to_numeric(df_dados[col].apply(clean_number_value), errors='coerce')

        if 'HVI' in df_dados.columns:
            indices_total = df_dados[df_dados['HVI'] == 'TOTAL'].index
            if not indices_total.empty: df_dados = df_dados.loc[:indices_total[0] - 1]
        
        df_dados = df_dados.dropna(axis=1, how='all')
        if not df_dados.empty: df_dados = df_dados.dropna(subset=[df_dados.columns[0]], how='all')
        if df_dados.empty: return None, None
        
        df_final = df_dados.copy()
        df_final.insert(0, 'CONT. REF', range(1, len(df_final) + 1))
        df_final.insert(0, 'Nº BUYER', n_buyer); df_final.insert(0, 'Nº AG', n_ag); df_final.insert(0, 'AGENT', agent)
        df_final.insert(0, 'BUYER', buyer); df_final.insert(0, 'SELLER', seller); df_final.insert(0, 'DATE', date_formatada)
        df_final.insert(0, 'NUMBER', number); df_final.insert(0, 'REF.CUCALA', ref_cucala)
        df_alinhado = df_final.reindex(columns=colunas_mestra)
        
        log(f"  -> Sucesso: {len(df_alinhado)} linhas extraídas e limpas.")
        return df_alinhado, (number, seller)
    except Exception as e:
        log(f"  -> ERRO GERAL ao processar o arquivo {os.path.basename(arquivo)}: {e}")
        log(traceback.format_exc())
        return None, None

def executar_logica_consolidacao(planilha_mestra, lista_arquivos_origem, log_queue, modo_atualizacao=False): # Função para executar a lógica de consolidação
    def log(msg):
        log_queue.put(msg)
    try:
        log("Lendo a planilha mestra...")
        df_mestra = pd.read_excel(planilha_mestra, header=1)
        df_mestra_limpa = df_mestra.dropna(subset=['CONT. REF'], how='all')
        colunas_mestra = df_mestra.columns.tolist()
    except Exception as e:
        log(f"ERRO: Não foi possível ler a planilha mestra: {e}")
        return False, "Falha ao ler planilha mestra."

    arquivos_processados = 0
    dados_para_adicionar = []
    df_mestra_modificada = df_mestra_limpa.copy()

    log("--- Iniciando verificação dos arquivos ---")
    for arquivo in lista_arquivos_origem:
        df_processado, id_lancamento = processar_arquivo_origem(arquivo, colunas_mestra, log_queue)
        
        if df_processado is None: continue

        number, seller = id_lancamento
        existe = not df_mestra_modificada[(df_mestra_modificada['NUMBER'] == number) & (df_mestra_modificada['SELLER'] == seller)].empty

        if modo_atualizacao:
            if existe:
                log(f"Lançamento '{number} - {seller}' encontrado. Marcando para atualização.")
                indices_para_remover = df_mestra_modificada[(df_mestra_modificada['NUMBER'] == number) & (df_mestra_modificada['SELLER'] == seller)].index
                df_mestra_modificada.drop(indices_para_remover, inplace=True)
                dados_para_adicionar.append(df_processado)
                arquivos_processados += 1
            else:
                log(f"AVISO: Lançamento '{number} - {seller}' não encontrado para atualizar.")
        else: # Modo Adição
            if not existe:
                log(f"Lançamento '{number} - {seller}' é novo. Marcando para adição.")
                dados_para_adicionar.append(df_processado)
                arquivos_processados += 1
            else:
                log(f"AVISO: Lançamento '{number} - {seller}' já existe e foi ignorado.")

    if arquivos_processados == 0:
        log("Nenhuma operação necessária.")
        return True, "Nenhuma operação necessária."

    df_final = pd.concat([df_mestra_modificada] + dados_para_adicionar, ignore_index=True)
    if 'UNIQUE ID' in df_final.columns:
        df_final['UNIQUE ID'] = range(1, len(df_final) + 1)

    msg_final = f"Operação concluída para {arquivos_processados} arquivo(s)."
    log(f"\n{msg_final}")
    
    try:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        backup_path = f"{os.path.splitext(planilha_mestra)[0]}_backup_{timestamp}.xlsx"
        log(f"Criando backup em: {os.path.basename(backup_path)}")
        shutil.copy(planilha_mestra, backup_path)

        workbook = load_workbook(planilha_mestra)
        sheet = workbook.active
        if sheet.max_row > 2:
            sheet.delete_rows(3, sheet.max_row)

        rows_to_add = dataframe_to_rows(df_final, index=False, header=False)
        log("Escrevendo novos dados na planilha...")
        for r_idx, row in enumerate(rows_to_add, 3):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        log("Salvando as alterações...")
        workbook.save(planilha_mestra)
        log(f"\n--- SUCESSO! A planilha foi atualizada preservando a formatação. ---")
        return True, msg_final
    except Exception as e:
        log(f"ERRO ao salvar os dados: {e} \n {traceback.format_exc()}")
        return False, "Falha ao salvar os dados."

# CLASSE DA JANELA DE LOGIN

class LoginWindow(ctk.CTkToplevel): # Classe da janela de login
    def __init__(self, master=None):
        super().__init__(master)
        self.title("Login - Consolidador CUCALA")
        self.geometry("350x420") 
        self.transient(master)
        self.grab_set()
        self.login_successful = False
        self.selected_theme = "dark" # Tema padrão

        self.VALID_CREDENTIALS = {"admin": "admin", "teste": "teste"} # Credenciais válidas
        self.grid_columnconfigure(0, weight=1)
        
        try:
            logo_path = resource_path("assets/logo.png")
            logo_image = ctk.CTkImage(Image.open(logo_path), size=(151, 151))
            logo_label = ctk.CTkLabel(self, image=logo_image, text="")
            logo_label.grid(row=0, column=0, pady=20)
        except FileNotFoundError:
            logo_label = ctk.CTkLabel(self, text="CUCALA", font=ctk.CTkFont(size=24, weight="bold"))
            logo_label.grid(row=0, column=0, pady=20)

        self.user_entry = ctk.CTkEntry(self, placeholder_text="Usuário", width=250)
        self.user_entry.grid(row=1, column=0, padx=30, pady=10)
        self.pass_entry = ctk.CTkEntry(self, placeholder_text="Senha", show="*", width=250)
        self.pass_entry.grid(row=2, column=0, padx=30, pady=10)
        self.pass_entry.bind("<Return>", self.check_login)
        self.login_button = ctk.CTkButton(self, text="Login", width=250, command=self.check_login)
        self.login_button.grid(row=3, column=0, padx=30, pady=15, ipady=5)
        self.status_label = ctk.CTkLabel(self, text="", text_color="red")
        self.status_label.grid(row=4, column=0, padx=30, pady=5)
        
        # Botão de tema na área de login que define qual tema será aplicado na interface principal 
        self.theme_switch = ctk.CTkSwitch(self, text="Dark", command=self.toggle_theme)
        self.theme_switch.grid(row=5, column=0, padx=30, pady=10)
        self.theme_switch.select()
    

    def check_login(self, event=None): # Verifica as credenciais inseridas
        user = self.user_entry.get()
        password = self.pass_entry.get()
        if user in self.VALID_CREDENTIALS and self.VALID_CREDENTIALS[user] == password:
            self.login_successful = True
            self.destroy()
        else:
            self.status_label.configure(text="Usuário ou senha inválida.")
    
    def toggle_theme(self): # Alterna o tema da janela de login
        if self.theme_switch.get() == 1:
            ctk.set_appearance_mode("dark")
            self.selected_theme = "dark"
            self.theme_switch.configure(text="Dark")
        else:
            ctk.set_appearance_mode("light")
            self.selected_theme = "light"
            self.theme_switch.configure(text="Light")

# CLASSE PRINCIPAL DA APLICAÇÃO (INTERFACE GRÁFICA)

class App(ctk.CTk):
    def __init__(self, initial_theme="light"):
        super().__init__()
        self.title("Consolidador de Planilhas CUCALA")
        self.geometry("1280x720")
        ctk.set_appearance_mode(initial_theme) # Define o tema com base na escolha da tela de login

        self.planilha_mestra_path = ""
        self.planilhas_origem_paths = []
        self.worker_thread = None
        self.log_queue = queue.Queue()
        
        self.grid_columnconfigure(0, weight=1); self.grid_columnconfigure(1, weight=2); self.grid_rowconfigure(1, weight=1)
        header = ctk.CTkFrame(self, fg_color="transparent"); header.grid(row=0, column=0, columnspan=2, padx=10, pady=10, sticky="ew"); header.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(header, text="Consolidador de Planilhas CUCALA v1.0", font=ctk.CTkFont(size=28, weight="bold")).grid(row=0, column=0)
        left_frame = ctk.CTkFrame(self); left_frame.grid(row=1, column=0, padx=(20, 10), pady=10, sticky="nsew"); left_frame.grid_columnconfigure(0, weight=1); left_frame.grid_rowconfigure(2, weight=1)
        try:
            logo_path = resource_path("assets/logo.png"); logo_image = ctk.CTkImage(Image.open(logo_path), size=(189, 189))
            ctk.CTkLabel(left_frame, image=logo_image, text="").grid(row=0, column=0, pady=20)
        except FileNotFoundError:
            ctk.CTkLabel(left_frame, text="Logo não encontrada", font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, column=0, pady=10)
        mestra_frame = ctk.CTkFrame(left_frame, fg_color="transparent"); mestra_frame.grid(row=1, column=0, padx=10, pady=10, sticky="ew"); mestra_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(mestra_frame, text="1. Selecione a Planilha Mestra:", font=ctk.CTkFont(size=16)).grid(row=0, column=0, sticky="w")
        self.entry_mestra = ctk.CTkEntry(mestra_frame, placeholder_text="Nenhum arquivo..."); self.entry_mestra.grid(row=1, column=0, pady=5, sticky="ew")
        ctk.CTkButton(mestra_frame, text="Procurar", width=80, command=self.selecionar_mestra).grid(row=1, column=1, padx=(10,0))
        origem_frame = ctk.CTkFrame(left_frame, fg_color="transparent"); origem_frame.grid(row=2, column=0, padx=10, pady=10, sticky="nsew"); origem_frame.grid_columnconfigure(0, weight=1); origem_frame.grid_rowconfigure(1, weight=1)
        ctk.CTkLabel(origem_frame, text="2. Selecione e Ordene as Planilhas:", font=ctk.CTkFont(size=16)).grid(row=0, column=0, sticky="w")
        listbox_container = ctk.CTkFrame(origem_frame); listbox_container.grid(row=1, column=0, pady=5, sticky="nsew"); listbox_container.grid_columnconfigure(0, weight=1); listbox_container.grid_rowconfigure(0, weight=1)
        self.listbox = tkinter.Listbox(listbox_container, background="#343638", foreground="white", selectbackground="#1F6AA5", borderwidth=0, highlightthickness=0, font=("Calibri", 12), selectmode=tkinter.EXTENDED)
        self.listbox.grid(row=0, column=0, padx=1, pady=1, sticky="nsew")
        reorder_frame = ctk.CTkFrame(origem_frame, fg_color="transparent"); reorder_frame.grid(row=2, column=0, pady=5, sticky="ew"); reorder_frame.grid_columnconfigure((0,1,2), weight=1)
        ctk.CTkButton(reorder_frame, text="Selecionar", command=self.selecionar_origem).grid(row=0, column=0, padx=(0,5), sticky="ew")
        ctk.CTkButton(reorder_frame, text="↑", command=self.move_up).grid(row=0, column=1, padx=5, sticky="ew")
        ctk.CTkButton(reorder_frame, text="↓", command=self.move_down).grid(row=0, column=2, padx=(5,0), sticky="ew")
        action_frame = ctk.CTkFrame(left_frame, fg_color="transparent"); action_frame.grid(row=3, column=0, padx=10, pady=20, sticky="ew");
        action_frame.grid_columnconfigure(0, weight=65); action_frame.grid_columnconfigure(1, weight=35)
        self.add_button = ctk.CTkButton(action_frame, text="Adicionar Novos Dados", height=50, font=ctk.CTkFont(size=18, weight="bold"), command=lambda: self.executar(modo_atualizacao=False))
        self.add_button.grid(row=0, column=0, ipady=10, sticky="ew", padx=(0,10))
        self.update_button = ctk.CTkButton(action_frame, text="Atualizar\nLançamento", height=50, font=ctk.CTkFont(size=14, weight="bold"), fg_color="#4A4D50", hover_color="#5F6266", command=lambda: self.executar(modo_atualizacao=True))
        self.update_button.grid(row=0, column=1, ipady=10, sticky="ew")
        right_frame = ctk.CTkFrame(self); right_frame.grid(row=1, column=1, padx=(10, 20), pady=10, sticky="nsew"); right_frame.grid_columnconfigure(0, weight=1); right_frame.grid_rowconfigure(1, weight=1)
        ctk.CTkLabel(right_frame, text="Caixa de diálogo do sistema:", font=ctk.CTkFont(size=16)).grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.log_textbox = ctk.CTkTextbox(right_frame, font=("Courier New", 12)); self.log_textbox.grid(row=1, column=0, padx=10, pady=(0,10), sticky="nsew")
        footer_frame = ctk.CTkFrame(self, fg_color="transparent"); footer_frame.grid(row=2, column=0, columnspan=2, padx=20, pady=10, sticky="ew"); footer_frame.grid_columnconfigure(0, weight=1); footer_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(footer_frame, text="Desenvolvido por Vinicios Reis para uso exclusivo CUCALA", text_color="gray", font=ctk.CTkFont(size=12)).grid(row=0, column=0, sticky="w")
        self.theme_switch = ctk.CTkSwitch(footer_frame, text="Dark", command=self.change_theme)
        self.theme_switch.grid(row=0, column=1, sticky="e")
        if initial_theme == "dark":
            self.theme_switch.select()
        self.change_theme() # Para definir as cores iniciais da listbox

    def change_theme(self): # Alterna o tema da interface
        is_dark = self.theme_switch.get() == 1
        new_mode = "dark" if is_dark else "light"
        ctk.set_appearance_mode(new_mode)
        
        if is_dark:
            self.theme_switch.configure(text="Dark")
            self.listbox.configure(background="#343638", foreground="white")
        else:
            self.theme_switch.configure(text="Light")
            self.listbox.configure(background="#EBEBEB", foreground="black")

    def selecionar_mestra(self):
        path = filedialog.askopenfilename(title="Selecione a Planilha Mestra", filetypes=[("Planilhas Excel", "*.xlsx")])
        if path: self.planilha_mestra_path = path; self.entry_mestra.delete(0, ctk.END); self.entry_mestra.insert(0, os.path.basename(path))

    def selecionar_origem(self):
        paths = filedialog.askopenfilenames(title="Selecione as Planilhas de Origem", filetypes=[("Planilhas Excel", "*.xlsx")])
        if paths:
            self.planilhas_origem_paths = list(paths); self.listbox.delete(0, ctk.END)
            for path in self.planilhas_origem_paths: self.listbox.insert(ctk.END, os.path.basename(path))

    def move_up(self):
        selected_indices = self.listbox.curselection()
        if not selected_indices: return
        for i in selected_indices:
            if i > 0:
                self.planilhas_origem_paths.insert(i-1, self.planilhas_origem_paths.pop(i)); text = self.listbox.get(i)
                self.listbox.delete(i); self.listbox.insert(i-1, text); self.listbox.selection_set(i-1)
    
    def move_down(self):
        selected_indices = self.listbox.curselection()
        if not selected_indices: return
        for i in reversed(selected_indices):
            if i < self.listbox.size() - 1:
                self.planilhas_origem_paths.insert(i+1, self.planilhas_origem_paths.pop(i)); text = self.listbox.get(i)
                self.listbox.delete(i); self.listbox.insert(i+1, text); self.listbox.selection_set(i+1)

    def process_queue(self): # Processa as mensagens do log
        try:
            while True:
                msg = self.log_queue.get_nowait()
                if isinstance(msg, tuple) and msg[0] == "WORKER_FINISHED":
                    self.finalize_process(msg[1])
                    return
                self.log_textbox.insert(ctk.END, msg + '\n')
                self.log_textbox.see(ctk.END)
        except queue.Empty:
            pass
        if self.worker_thread and self.worker_thread.is_alive():
            self.after(100, self.process_queue)

    def executar(self, modo_atualizacao): # Executa o processo de união das planilhas   
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showwarning("Aguarde", "Um processo já está em andamento.")
            return
        self.log_textbox.delete("1.0", ctk.END)
        if not self.planilha_mestra_path or not self.planilhas_origem_paths:
            messagebox.showerror("Erro de Validação", "Selecione a planilha mestra e as de origem.")
            return
        
        arquivos_para_processar = []
        if modo_atualizacao:
            selected_indices = self.listbox.curselection()
            if len(selected_indices) != 1:
                messagebox.showerror("Seleção Inválida", "Para atualizar, por favor, selecione apenas UM lançamento na lista.")
                return
            arquivos_para_processar = [self.planilhas_origem_paths[selected_indices[0]]]
        else:
            arquivos_para_processar = self.planilhas_origem_paths

        self.add_button.configure(state="disabled")
        self.update_button.configure(state="disabled")
        
        self.worker_thread = threading.Thread(target=self.run_consolidation_worker, args=(arquivos_para_processar, modo_atualizacao))
        self.worker_thread.start()
        self.after(100, self.process_queue)

    def run_consolidation_worker(self, arquivos_para_processar, modo_atualizacao): # Executa o processo de união das planilhas
        try:
            sucesso, msg_final = executar_logica_consolidacao(self.planilha_mestra_path, arquivos_para_processar, self.log_queue, modo_atualizacao)
            self.log_queue.put(("WORKER_FINISHED", (sucesso, msg_final)))
        except Exception as e:
            tb_str = traceback.format_exc()
            self.log_queue.put(f"ERRO CRÍTICO NO WORKER: {e}\n{tb_str}")
            self.log_queue.put(("WORKER_FINISHED", (False, f"Erro crítico: {e}")))
            
    def finalize_process(self, result): # Finaliza o processo de união das planilhas
        sucesso, msg_final = result
        self.add_button.configure(state="normal")
        self.update_button.configure(state="normal")
        if sucesso:
            messagebox.showinfo("Sucesso", f"Processo concluído!\n\n{msg_final}")
        else:
            messagebox.showerror("Falha na Execução", f"O processo falhou.\n\nDetalhe: {msg_final}")

# PONTO DE ENTRADA DO PROGRAMA

if __name__ == "__main__":
    
    ctk.set_appearance_mode("dark") # Define um tema inicial padrão, caso a janela de login seja fechada
    
    root = ctk.CTk()
    root.withdraw()

    login_window = LoginWindow(root)
    root.wait_window(login_window)

    if login_window.login_successful: # Pega o tema escolhido na janela de login
        selected_theme = login_window.selected_theme
        root.destroy()
        app = App(initial_theme=selected_theme) # Inicia a aplicação principal com o tema escolhido
        app.mainloop()
    else:
        root.destroy()